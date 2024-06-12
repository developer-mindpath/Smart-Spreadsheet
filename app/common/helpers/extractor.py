from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell


class ExcelDataWizard:

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=True)
        self.sheet = self.wb.active
        self.tables = []

    def check_border(self, cell: Cell) -> bool:
        border = cell.border
        return any([
            border.left.style,
            border.right.style,
            border.top.style,
            border.bottom.style
        ])

    def retrieve_tables(self):
        rows = list(self.sheet.iter_rows())
        table_markers = np.zeros((len(rows), len(rows[0])), dtype=bool)

        def retrieve_table(start_row: int, start_col: int) -> list[list[Any]]:
            end_col = start_col
            while end_col < len(rows[start_row]) and self.check_border(rows[start_row][end_col]):
                end_col += 1

            table = []
            for row_idx in range(start_row, len(rows)):
                if not self.check_border(rows[row_idx][start_col]):
                    break
                row_data = [
                    rows[row_idx][col_idx].value
                    for col_idx in range(start_col, end_col)
                ]
                table_markers[row_idx, start_col:end_col] = True
                table.append(row_data)

            return table

        for row_idx, row in enumerate(rows):
            for col_idx, cell in enumerate(row):
                if not table_markers[row_idx, col_idx] and self.check_border(cell):
                    table = retrieve_table(row_idx, col_idx)
                    self.tables.append(table)

    def convert_to_dfs(self) -> list[pd.DataFrame]:
        return [pd.DataFrame(table) for table in self.tables]
